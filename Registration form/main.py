# Import modules

from openpyxl import *
from tkinter import *

# Declaring workbook and sheet variables

workbook = load_workbook("C:\\Users\\heckt\\OneDrive\\Desktop\\users.xlsx")

# Creating sheet object

sheet = workbook.active

def excel ():
    # Resize the width of columns in excel spreadsheet

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # Write given data at particular locations

    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "Course"
    sheet.cell(row = 1, column = 3).value = "Semester"
    sheet.cell(row = 1, column = 4).value = "Form Number"
    sheet.cell(row = 1, column = 5).value = "Contact Number"
    sheet.cell(row = 1, column = 6).value = "Email"
    sheet.cell(row = 1, column = 7).value = "Address"

# Function to set cursor's focus

def first_focus (event):
    # Set focus on the course_field box

    course_field.focus_set()

def second_focus (event):
    # Set focus on the sem_field box

    sem_field.focus_set()

def third_focus (event):
    # Set focus on the form_no_field box

    form_no_field.focus_set()

def fourth_focus (event):
    # Set focus on the contact_no_field box

    contact_no_field.focus_set()

def fifth_focus (event):
    # Set focus on the email_field box

    email_field.focus_set()

def sixth_focus (event):
    # Set focus on the address_field box

    address_field.focus_set()

def clear ():
    # Clear content of entry box

    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_field.delete(0, END)
    address_field.delete(0, END)

# Function to take data from GUI window and write it into workbook

def insert ():
    # If user didn't fill any entry

    if (name_field.get() == "" and course_field.get() == ""
        and sem_field.get() == "" and form_no_field.get() == ""
        and contact_no_field.get() == "" and email_field.get()
        == "" and address_field.get() == ""):
        print("Empty input")
    else:
        # Assigning max row and max column value into Excel

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        workbook.save("C:\\Users\\heckt\\OneDrive\\Desktop\\users.xlsx")

        # Set focus on the name_field box

        name_field.focus_set()

        clear()


root = Tk()

root.configure(background = "light green")

root.title("Registration form")

root.geometry("500x300")

excel()

heading = Label(root, text="Form", bg="light green")

name = Label(root, text="Name", bg="light green")

course = Label(root, text="Course", bg="light green")

sem = Label(root, text="Semester", bg="light green")

form_no = Label(root, text="Form No.", bg="light green")

contact_no = Label(root, text="Contact No.", bg="light green")

email = Label(root, text="Email", bg="light green")

address = Label(root, text="Address", bg="light green")

# Grid method for placing items

heading.grid(row = 0, column = 1)
name.grid(row = 1, column = 0)
course.grid(row = 2, column = 0)
sem.grid(row = 3, column = 0)
form_no.grid(row = 4, column = 0)
contact_no.grid(row = 5, column = 0)
email.grid(row = 6, column = 0)
address.grid(row = 7, column = 0)

# Creating entry boxes

name_field = Entry(root)
course_field = Entry(root)
sem_field = Entry(root)
form_no_field = Entry(root)
contact_no_field = Entry(root)
email_field = Entry(root)
address_field = Entry(root)

# Binding methods of widgets

# Whenever "Enter" is pressed, we are calling "focus" functions

name_field.bind("<Return>", first_focus)
course_field.bind("<Return>", second_focus)
sem_field.bind("<Return>", third_focus)
form_no_field.bind("<Return>", fourth_focus)
contact_no_field.bind("<Return>", fifth_focus)
email_field.bind("<Return>", sixth_focus)

# Grid method for placing widgets

name_field.grid(row = 1, column = 1, ipadx = "100")
course_field.grid(row = 2, column = 1, ipadx = "100")
sem_field.grid(row = 3, column = 1, ipadx = "100")
form_no_field.grid(row = 4, column = 1, ipadx = "100")
contact_no_field.grid(row = 5, column = 1, ipadx = "100")
email_field.grid(row = 6, column = 1, ipadx = "100")
address_field.grid(row = 7, column = 1, ipadx = "100")

excel()

submit = Button(root, text = "Submit", fg = "White", bg = "Black", command = insert)
submit.grid(row = 8, column = 1)

root.mainloop()
